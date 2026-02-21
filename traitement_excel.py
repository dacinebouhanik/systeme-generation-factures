from os import listdir, system, chdir
import os
import openpyxl
import datetime
from operator import itemgetter


###### Choix répértoire #######
def folderChoice():
    print("#################### CHOIX DU REPERTOIRE ###################")
    print
    print("Entrez le répertoire qui contient les fichiers xlsx (tableur excel au format xlsx)")
    workFolder = input("Veuillez entrez l'arboressence ex:  c:\mondossier :  ")
    return workFolder


def folderconvert(workFolder):
    currentFolder = ""
    for x in workFolder:
        if x == "\\":
            currentFolder = currentFolder + "/"
        else:
            currentFolder = currentFolder + x
    return currentFolder


def erreur():
    os.system('cls')
    print()
    print("Il y a une erreur ce dossier ne contient pas que des fichiers xlsx ou contient des dossier également.")
    print("Veuillez mettre un repertoire ne contenant QUE les fichiers XLSX")
    print("ou les fichiers XLSX présent ne sont pas conforme.")
    print(
        "Si vous avez déjà executé ce programme dans ce dossier il est probable que le fichier généré factureclients soit crée, il faut alors le supprimé.")
    input("Appuyez sur une touche pour terminer le programme.")
    exit()


def filetest(fileInFolderList):
    filelist = []
    for file in fileInFolderList:
        ext = file[-4:]
        if ext == "xlsx":
            filelist.append(file)
        elif ext == "lsx#":
            continue
        else:
            erreur()

    print("Test de fichier terminé... Le dossier contient que les fichiers xlsx....")
    print()
    return filelist


def checkfile(checkfilelist):
    for file in checkfilelist:
        print("les fichiers testé sont : " + str(file))
        wb = openpyxl.load_workbook(file, data_only=True)
        print(wb.sheetnames)
        sheet = wb["Feuille1"]
        cell = sheet["D1"]
        if not cell.value == 'feuille de commande': erreur(); return
        cell = sheet["A7"]
        if not cell.value == 'clients': erreur(); return
        cell = sheet["B7"]
        if not cell.value == 'maxi': erreur(); return
        cell = sheet["C7"]
        if not cell.value == 'geant': erreur(); return
        cell = sheet["D7"]
        if not cell.value == 'miche': erreur(); return
        print("Les fichiers XLSX sont conforme.")
    return


def takeclientlist(filelist):
    nombreclient = []
    for file in filelist:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb["Feuille1"]
        print("\n" + file)
        x = 8
        while True:
            cell = sheet.cell(x, 1)
            if not cell.value:
                break
            else:
                print(cell.value)
                nombreclient.append(cell.value)
                x = x + 1
    nombreclient = list(set(nombreclient))
    return nombreclient


def getcustomersdata(allCustomersInFolderList, filelist):
    totalCustomersData = []
    customersData = []
    for customers in allCustomersInFolderList:
        customersData = []
        for file in filelist:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb["Feuille1"]
            print("\n" + file)
            x = 7
            while True:
                cell = sheet.cell(x, 1)
                if not cell.value:
                    break
                else:
                    if (cell.value) == customers:
                        customersData.append(
                            [customers, sheet["B4"].value, sheet["B" + str(x)].value, sheet["C" + str(x)].value,
                             sheet["D" + str(x)].value, sheet["E" + str(x)].value, sheet["F" + str(x)].value,
                             sheet["G" + str(x)].value, sheet["H" + str(x)].value])
                        x = 7
                        break
                    else:
                        x = x + 1
        totalCustomersData.append(customersData)
    return totalCustomersData